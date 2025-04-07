from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from bs4 import BeautifulSoup
import os
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from datetime import datetime  # Import datetime module


@csrf_exempt  # Allow POST requests without CSRF token for testing (remove in production)
def report_xlsx(request):
    if request.method == 'POST':
        # Get form data
        from_date = request.POST.get('from_date', '')
        to_date = request.POST.get('to_date', '')
        mode = request.POST.get('mode', '')
        part_model = request.POST.get('part_model', '')
        shift = request.POST.get('shift', '')
        status = request.POST.get('status', '')
        total_count = request.POST.get('total_count', '')
        # Get the table HTML
        table_html = request.POST.get('table_html')

        if table_html:
            # Parse the table HTML to extract data
            soup = BeautifulSoup(table_html, 'html.parser')
            thead = soup.find('thead')
            rows = soup.find_all('tr')

            # Create a new Excel workbook
            wb = Workbook()
            ws = wb.active

            # Set headers for A1 to A4 with spacing
            ws['A1'] = f'From Date: {from_date}'
            ws['B1'] = f'To Date: {to_date}'
            ws['D1'] = f'Mode: {mode}'
            ws['C1'] = f'Part Model: {part_model}'
            ws['E1'] = f'Shift: {shift}'
            ws['G1'] = f'Status: {status}'
            ws['F1'] = f'Total_count: {total_count}'

            # Apply bold and alignment styles
            for cell in ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1']:
                ws[cell].font = Font(bold=True)
                ws[cell].alignment = Alignment(horizontal='left')

            # Write the table headers (thead) starting at A5
            start_row = 5
            if thead:
                header_cells = thead.find_all('th')
                for col_num, th in enumerate(header_cells, start=1):
                    cell = ws.cell(row=start_row, column=col_num, value=th.text.strip())
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')

            # Write table rows (tr) starting from A6
            for row_num, row in enumerate(rows[1:], start=start_row + 1):  # Skip the header row
                cells = row.find_all(['td', 'th'])
                for col_num, cell in enumerate(cells, start=1):
                    ws.cell(row=row_num, column=col_num, value=cell.text.strip())

            # Adjust column widths based on max content length
            for col_num in range(1, ws.max_column + 1):
                max_length = 0
                col_letter = get_column_letter(col_num)
                for cell in ws[col_letter]:
                    try:
                        max_length = max(max_length, len(str(cell.value)) if cell.value else 0)
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2  # Add padding

            current_datetime = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')

            # Generate filename using the current date and time
            file_name = f'report_data_{current_datetime}.xlsx'
            
            
            # Save the XLSX file to Downloads directory with dynamic name
            downloads_dir = r'C:\Program Files\Four_Channel_Rasperripi\xlsx_files'
            file_path = os.path.join(downloads_dir, file_name)
            wb.save(file_path)

            # Return success response
            return JsonResponse({'success': True, 'message': f'File saved at: {file_path}'})
        else:
            return JsonResponse({'success': False, 'message': 'No table data provided.'})
    else:
        return JsonResponse({'success': False, 'message': 'Invalid request method.'})
